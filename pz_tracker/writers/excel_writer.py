"""Write extracted P&Z data into the Excel workbook using openpyxl."""

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_COLUMNS, EXCEL_FILE
from utils.deduplicator import entry_exists

logger = logging.getLogger(__name__)


def _get_or_create_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    """Return the worksheet named *sheet_name*, creating it if needed.

    New sheets are initialised with the standard header row.
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    for col_idx, header in enumerate(EXCEL_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    logger.info("Created new sheet: %s", sheet_name)
    return ws


def _load_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    # Remove the default empty sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def write_records(
    records: list[dict],
    city: str,
    excel_path: Path | None = None,
) -> int:
    """Append *records* to the city's sheet, skipping duplicates.

    Parameters
    ----------
    records : list[dict]
        Each dict should contain keys matching the AI extractor output:
        applicant_name, applicant_business_name, applicant_email,
        applicant_phone, construction_type, land_acres, location,
        description, pz_meeting_date.  Plus optional ``url`` and ``status``.
    city : str
        City name — used as the sheet tab name and the "City" column value.
    excel_path : Path, optional
        Override for the default EXCEL_FILE path.

    Returns
    -------
    int
        Number of new rows actually written (after dedup).
    """
    excel_path = excel_path or EXCEL_FILE
    wb = _load_workbook(excel_path)
    ws = _get_or_create_sheet(wb, city)

    written = 0
    for rec in records:
        url = rec.get("url", "")
        description = rec.get("description", "")

        if entry_exists(ws, url, description):
            logger.debug("Skipping duplicate: %s", url or description[:60])
            continue

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=city)
        ws.cell(row=next_row, column=2, value=rec.get("pz_meeting_date", ""))
        ws.cell(row=next_row, column=3, value=rec.get("owner_name", ""))
        ws.cell(row=next_row, column=4, value=rec.get("general_contractor", ""))
        ws.cell(row=next_row, column=5, value=rec.get("architect", ""))
        ws.cell(row=next_row, column=6, value=rec.get("applicant_name", ""))
        ws.cell(row=next_row, column=7, value=rec.get("applicant_email", ""))
        ws.cell(row=next_row, column=8, value=rec.get("applicant_phone", ""))
        ws.cell(row=next_row, column=9, value=rec.get("construction_type", ""))
        ws.cell(row=next_row, column=10, value=rec.get("land_acres", ""))
        ws.cell(row=next_row, column=11, value=rec.get("location", ""))
        ws.cell(row=next_row, column=12, value=description)
        ws.cell(row=next_row, column=13, value=url)
        written += 1

    wb.save(excel_path)
    logger.info("Wrote %d new record(s) to sheet '%s' in %s", written, city, excel_path.name)
    return written


def write_no_data_record(
    city: str,
    reason: str = "No relevant data found",
    excel_path: Path | None = None,
) -> None:
    """Write a single row indicating no relevant data was found for *city*."""
    excel_path = excel_path or EXCEL_FILE
    wb = _load_workbook(excel_path)
    ws = _get_or_create_sheet(wb, city)

    # Check if a "no data" row already exists
    for row_num in range(2, ws.max_row + 1):
        desc = ws.cell(row=row_num, column=12).value or ""
        if desc.startswith("No relevant data"):
            return  # already recorded

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=city)
    ws.cell(row=next_row, column=12, value=reason)
    wb.save(excel_path)
    logger.info("Recorded no-data for '%s' in %s", city, excel_path.name)
