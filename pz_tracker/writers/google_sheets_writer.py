"""Write extracted P&Z data into a Google Sheets spreadsheet via gspread.

Data is organised by **portal type** – one tab per portal (e.g. "CivicPlus",
"MuniCode", "CivicClerk", …).  Each row carries the city name in the first
column so all cities sharing a portal live on the same sheet.
"""

import json
import logging
import os
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials

from config import DEFAULT_CITIES, EXCEL_COLUMNS

logger = logging.getLogger(__name__)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

SPREADSHEET_ID = "1rwCdWEp1g713XMYaZY_3xpUHnm0Ynlujuua2FoHoOGo"

# Pretty names for the sheet tabs
_PORTAL_TAB_NAMES: dict[str, str] = {
    "civicplus": "CivicPlus",
    "municode": "MuniCode",
    "civicclerk": "CivicClerk",
    "civicweb": "CivicWeb",
    "legistar": "Legistar",
    "standard_html": "Standard HTML",
}


def _get_client() -> gspread.Client:
    """Authenticate with Google using service account credentials."""
    creds_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
    if creds_path and Path(creds_path).exists():
        creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    else:
        # Try loading from env variable as raw JSON
        creds_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_CREDS")
        if creds_json:
            info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(info, scopes=SCOPES)
        else:
            raise RuntimeError(
                "Google credentials not found. Set GOOGLE_SERVICE_ACCOUNT_JSON "
                "(path to JSON file) or GOOGLE_SERVICE_ACCOUNT_CREDS (raw JSON) "
                "in your .env file."
            )
    return gspread.authorize(creds)


def _get_or_create_sheet(
    spreadsheet: gspread.Spreadsheet, sheet_name: str
) -> gspread.Worksheet:
    """Return the worksheet named *sheet_name*, creating it if needed."""
    try:
        ws = spreadsheet.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=sheet_name, rows=100, cols=len(EXCEL_COLUMNS))
        ws.append_row(EXCEL_COLUMNS, value_input_option="RAW")
        logger.info("Created new sheet tab: %s", sheet_name)
    return ws


def _get_existing_descriptions(ws: gspread.Worksheet) -> set[str]:
    """Return a set of (url, description) tuples already in the sheet for dedup."""
    rows = ws.get_all_values()
    existing = set()
    for row in rows[1:]:  # skip header
        if len(row) >= 13:
            url = row[12].strip() if row[12] else ""
            desc = row[11].strip() if row[11] else ""
            existing.add((url, desc))
    return existing


def write_records_to_sheets(
    records: list[dict],
    city: str,
    portal_type: str = "",
    spreadsheet_id: str | None = None,
) -> int:
    """Append *records* to the portal-type sheet tab in Google Sheets.

    Returns the number of new rows written (after dedup).
    """
    spreadsheet_id = spreadsheet_id or SPREADSHEET_ID
    tab_name = _PORTAL_TAB_NAMES.get(portal_type, portal_type or city)

    client = _get_client()
    spreadsheet = client.open_by_key(spreadsheet_id)
    ws = _get_or_create_sheet(spreadsheet, tab_name)

    existing = _get_existing_descriptions(ws)
    written = 0

    rows_to_append = []
    for rec in records:
        url = rec.get("url") or ""
        description = rec.get("description") or ""
        if (url.strip(), description.strip()) in existing:
            logger.debug("Skipping duplicate in Sheets: %s", url or description[:60])
            continue

        row = [
            city,
            rec.get("pz_meeting_date", ""),
            rec.get("owner_name", ""),
            rec.get("general_contractor", ""),
            rec.get("architect", ""),
            rec.get("applicant_name", ""),
            rec.get("applicant_email", ""),
            rec.get("applicant_phone", ""),
            rec.get("construction_type", ""),
            rec.get("land_acres", ""),
            rec.get("location", ""),
            description,
            url,
        ]
        rows_to_append.append(row)
        written += 1

    if rows_to_append:
        ws.append_rows(rows_to_append, value_input_option="RAW")

    logger.info(
        "Wrote %d new record(s) to Google Sheet tab '%s' (%s)", written, tab_name, city
    )
    return written


def write_no_data_to_sheets(
    city: str,
    portal_type: str = "",
    reason: str = "No relevant data found",
    spreadsheet_id: str | None = None,
) -> None:
    """Write a single row indicating no relevant data was found for *city*."""
    spreadsheet_id = spreadsheet_id or SPREADSHEET_ID
    tab_name = _PORTAL_TAB_NAMES.get(portal_type, portal_type or city)

    client = _get_client()
    spreadsheet = client.open_by_key(spreadsheet_id)
    ws = _get_or_create_sheet(spreadsheet, tab_name)

    # Check if already recorded
    existing = ws.get_all_values()
    for row in existing[1:]:
        if len(row) >= 12 and row[0] == city and row[11].startswith("No relevant data"):
            return  # already recorded

    row = [city, "", "", "", "", "", "", "", "", "", "", reason, ""]
    ws.append_row(row, value_input_option="RAW")
    logger.info("Recorded no-data for '%s' in Google Sheet tab '%s'", city, tab_name)


def copy_excel_to_sheets(
    excel_path: Path,
    spreadsheet_id: str | None = None,
) -> dict[str, int]:
    """Copy all data from the local Excel workbook into Google Sheets.

    Rows are grouped by **portal type** (one tab per portal).  The city
    name in each row is looked up in ``DEFAULT_CITIES`` to determine
    the portal type.

    Returns {tab_name: rows_written}.
    """
    from openpyxl import load_workbook

    spreadsheet_id = spreadsheet_id or SPREADSHEET_ID
    client = _get_client()
    spreadsheet = client.open_by_key(spreadsheet_id)
    wb = load_workbook(excel_path)

    # Collect all Excel rows grouped by portal type
    portal_rows: dict[str, list[list[str]]] = {}

    for sheet_name in wb.sheetnames:
        ws_xl = wb[sheet_name]
        # Determine portal from city config, fall back to sheet name itself
        city_info = DEFAULT_CITIES.get(sheet_name, {})
        portal_type = city_info.get("portal_type", "")
        tab_name = _PORTAL_TAB_NAMES.get(portal_type, portal_type or sheet_name)

        for row_num in range(2, ws_xl.max_row + 1):  # skip header
            row = []
            for col in range(1, len(EXCEL_COLUMNS) + 1):
                val = ws_xl.cell(row=row_num, column=col).value
                row.append(str(val) if val is not None else "")
            portal_rows.setdefault(tab_name, []).append(row)

    summary: dict[str, int] = {}

    for tab_name, rows in portal_rows.items():
        ws_gs = _get_or_create_sheet(spreadsheet, tab_name)
        existing = _get_existing_descriptions(ws_gs)

        rows_to_append = []
        for row in rows:
            url = row[12].strip() if len(row) > 12 else ""
            desc = row[11].strip() if len(row) > 11 else ""
            if (url, desc) in existing:
                continue
            rows_to_append.append(row)

        if rows_to_append:
            ws_gs.append_rows(rows_to_append, value_input_option="RAW")

        summary[tab_name] = len(rows_to_append)
        logger.info(
            "Copied %d row(s) to Google Sheet tab '%s'",
            len(rows_to_append),
            tab_name,
        )

    return summary
