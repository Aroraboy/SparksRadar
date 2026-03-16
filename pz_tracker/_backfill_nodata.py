"""Backfill 'no data found' entries for cities already processed with zero records."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from dotenv import load_dotenv
load_dotenv()

import logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s")

from openpyxl import load_workbook
from config import DEFAULT_CITIES, EXCEL_FILE
from writers.excel_writer import write_no_data_record
from writers.google_sheets_writer import write_no_data_to_sheets

# Cities from Phase 1 (CivicClerk) and Phase 2 (CivicPlus) that returned 0 records
NO_DATA_CITIES = [
    # Phase 1 CivicClerk - no data
    "Alvarado", "Kaufman", "Sachse",
    # Phase 1 CivicClerk - API 404
    "Belton", "Addison", "Josephine",
    # Phase 2 CivicPlus - no data
    "Melissa", "Balch Springs", "Wilmer", "Royse City", "Crandall",
    "Seagoville", "Ovilla", "Frisco", "Gunter", "Anna",
    "The Colony", "Lancaster", "Liberty Hill", "Cedar Hill",
    "Sunnyvale", "Corinth",
]

# Check which cities already have data sheets in Excel
wb = load_workbook(EXCEL_FILE)
cities_with_data = set()
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # If it has more than just headers or no-data rows, it has real data
    has_real_data = False
    for row_num in range(2, ws.max_row + 1):
        desc = ws.cell(row=row_num, column=12).value or ""
        if not desc.startswith("No relevant data"):
            has_real_data = True
            break
    if has_real_data:
        cities_with_data.add(sheet_name)
wb.close()

print(f"Cities with real data in Excel: {len(cities_with_data)}")

import time

for city in NO_DATA_CITIES:
    if city in cities_with_data:
        print(f"  SKIP {city} (has real data)")
        continue

    info = DEFAULT_CITIES.get(city, {})
    portal_type = info.get("portal_type", "")

    if city in ("Belton", "Addison", "Josephine"):
        reason = "No relevant data found – CivicClerk API unavailable"
    elif city in ("Crandall", "Sunnyvale"):
        reason = "No relevant data found – website timeout"
    elif city in ("Balch Springs", "Seagoville", "Liberty Hill", "Cedar Hill"):
        reason = f"No relevant data found – latest agenda too old"
    elif city in ("Frisco", "Anna", "The Colony", "Lancaster", "Corinth"):
        reason = "No relevant data found – no agenda located"
    elif city == "Wilmer":
        reason = "No relevant data found – no P&Z agenda found"
    else:
        reason = "No relevant data found – agenda had no matching items"

    print(f"  Recording no-data: {city} ({portal_type})")
    write_no_data_record(city, reason=reason, excel_path=EXCEL_FILE)

    try:
        write_no_data_to_sheets(city, portal_type=portal_type, reason=reason)
        time.sleep(1.5)  # respect Sheets API rate limits
    except Exception as e:
        print(f"    Google Sheets error: {e}")

print("Done!")
