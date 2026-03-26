"""Inspect CC permit XLSX columns and data."""
import httpx
import io
import openpyxl

c = httpx.Client(follow_redirects=True, timeout=30, headers={"User-Agent": "Mozilla/5.0"})

# Permit report - all columns
r = c.get("https://www.corpuschristitx.gov/media/ba1nbcm4/permit-report-jan-2026-excel-version.xlsx")
wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
ws = wb[wb.sheetnames[0]]

# Get header row (row 8)
headers = [cell.value for cell in ws[8]]
print("PERMIT REPORT HEADERS:")
for i, h in enumerate(headers):
    if h:
        print(f"  Col {i}: {h}")

# Print some data rows
print("\nSAMPLE DATA ROWS:")
for row_idx in range(9, min(14, ws.max_row + 1)):
    vals = [cell.value for cell in ws[row_idx]]
    row_dict = {headers[i]: vals[i] for i in range(len(headers)) if headers[i] and vals[i] is not None}
    print(f"  Row {row_idx}: {row_dict}")

# CO report - all columns
r = c.get("https://www.corpuschristitx.gov/media/riplqktg/coo-report-jan-2026-excel-version.xlsx")
wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
ws = wb[wb.sheetnames[0]]

headers = [cell.value for cell in ws[8]]
print("\n\nCO REPORT HEADERS:")
for i, h in enumerate(headers):
    if h:
        print(f"  Col {i}: {h}")

print("\nCO SAMPLE DATA:")
for row_idx in range(9, min(14, ws.max_row + 1)):
    vals = [cell.value for cell in ws[row_idx]]
    row_dict = {headers[i]: vals[i] for i in range(len(headers)) if headers[i] and vals[i] is not None}
    print(f"  Row {row_idx}: {row_dict}")

c.close()
print("\nDone.")
