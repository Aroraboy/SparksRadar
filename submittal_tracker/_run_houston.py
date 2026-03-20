"""One-shot script to download Houston 2026 Plat Activity Reports and write to Google Sheet."""

import datetime
import io

import gspread
import httpx
import openpyxl
from google.oauth2.service_account import Credentials

# --- Houston 2026 Plat Activity Report URLs ---
REPORT_URLS = [
    ("January 8, 2026", "https://www.houstontx.gov/planning/DevelopRegs/docs_pdfs/Plat_report/2026/Plat-Current-Agenda-Spreadsheet-01-05-2026.xlsx"),
    ("January 22, 2026", "https://www.houstontx.gov/planning/DevelopRegs/docs_pdfs/Plat_report/2026/Plat-Current-Agenda-Spreadsheet-01-22-2026.xlsx"),
    ("February 5, 2026", "https://www.houstontx.gov/planning/DevelopRegs/docs_pdfs/Plat_report/2026/Plat-Current-Agenda-Spreadsheet-02-05-2026.xlsx"),
    ("March 5, 2026", "https://www.houstontx.gov/planning/DevelopRegs/docs_pdfs/Plat_report/2026/Plat-Current-Agenda-Spreadsheet-03-05-2026.xlsx"),
    ("March 19, 2026", "https://www.houstontx.gov/planning/DevelopRegs/docs_pdfs/Plat_report/2026/Plat-Current-Agenda-Spreadsheet-03-19-2026.xlsx"),
]

SPREADSHEET_ID = "14qiDFhK9BIsGDMnRMuVxkpfmfYqNpg5w-6nb48M5WaM"

SHEET_COLUMNS = [
    "Case Number", "Type", "Address", "Description",
    "Status", "Submittal Date", "Notes", "Source File", "Lead Priority",
]

HIGH_KW = [
    "warehouse", "distribution center", "office building", "mixed use",
    "mixed-use", "hospital", "medical center", "hotel", "apartment",
    "multifamily", "multi-family", "shopping center", "retail center",
    "commercial", "industrial", "manufacturing",
]
MEDIUM_KW = [
    "restaurant", "retail", "office", "church", "school", "daycare",
    "day care", "clinic", "garage", "parking", "gas station", "fuel",
    "car wash", "storage", "self-storage", "assisted living", "senior living",
]


def derive_priority(land_use: str, name: str) -> str:
    text = f"{land_use or ''} {name or ''}".lower()
    for kw in HIGH_KW:
        if kw in text:
            return "High"
    for kw in MEDIUM_KW:
        if kw in text:
            return "Medium"
    return "Low"


def fmt_date(dt) -> str:
    if isinstance(dt, datetime.datetime):
        return dt.strftime("%Y-%m-%d")
    return str(dt) if dt else ""


def safe(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def main():
    all_rows: list[list[str]] = []
    seen_app_nos: set[str] = set()
    client = httpx.Client(follow_redirects=True, timeout=30)

    for label, url in REPORT_URLS:
        resp = client.get(url)
        if resp.status_code != 200:
            print(f"SKIP {label}: HTTP {resp.status_code}")
            continue
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        count = 0
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row_cells[0]:
                continue
            app_no = safe(row_cells[1])
            if app_no in seen_app_nos:
                continue
            seen_app_nos.add(app_no)

            subdivision_name = safe(row_cells[0])
            app_type = safe(row_cells[4])
            land_use = safe(row_cells[21])
            acreage = row_cells[22]
            developer = safe(row_cells[25])
            pc_date = row_cells[2]
            county = safe(row_cells[9])
            zipcode = safe(row_cells[13])
            in_city = safe(row_cells[8])
            major_road = safe(row_cells[20])

            # Build address from location info
            addr_parts: list[str] = []
            if major_road:
                addr_parts.append(major_road)
            if county:
                addr_parts.append(f"{county} County")
            if zipcode:
                addr_parts.append(zipcode)
            if in_city:
                addr_parts.append(f"({in_city})")
            address = ", ".join(addr_parts)

            # Build description
            desc_parts: list[str] = []
            if subdivision_name:
                desc_parts.append(subdivision_name)
            if land_use:
                desc_parts.append(land_use)
            if acreage:
                desc_parts.append(f"{acreage} acres")
            description = "; ".join(desc_parts)

            priority = derive_priority(land_use, subdivision_name)

            notes = f"Developer: {developer}" if developer else ""

            all_rows.append([
                app_no,
                app_type,
                address,
                description,
                "",
                fmt_date(pc_date),
                notes,
                f"Plat Activity Report - {label}",
                priority,
            ])
            count += 1
        print(f"{label}: {count} rows extracted")

    print(f"Total unique rows: {len(all_rows)}")

    # --- Write to Google Sheets ---
    creds = Credentials.from_service_account_file(
        "pz_tracker/service_account.json",
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(SPREADSHEET_ID)

    # Get or create Houston tab
    try:
        sheet = ss.worksheet("Houston")
        first_row = sheet.row_values(1)
        if not first_row:
            sheet.append_row(SHEET_COLUMNS, value_input_option="RAW")
    except gspread.WorksheetNotFound:
        sheet = ss.add_worksheet(title="Houston", rows=800, cols=len(SHEET_COLUMNS))
        sheet.append_row(SHEET_COLUMNS, value_input_option="RAW")
        print("Created Houston tab")

    # Check existing to avoid dupes
    existing: set[tuple[str, str]] = set()
    for row in sheet.get_all_values()[1:]:
        if len(row) >= 3:
            existing.add((row[0].strip(), row[2].strip()))

    new_rows: list[list[str]] = []
    for r in all_rows:
        key = (r[0].strip(), r[2].strip())
        if key not in existing:
            existing.add(key)
            new_rows.append(r)

    if new_rows:
        sheet.append_rows(new_rows, value_input_option="RAW")

    print(f"Wrote {len(new_rows)} new rows to Houston tab ({len(all_rows) - len(new_rows)} duplicates skipped)")


if __name__ == "__main__":
    main()
