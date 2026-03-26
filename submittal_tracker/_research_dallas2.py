"""Dallas - check the new Building Permits dataset and the monthly XLSX reports."""
import httpx

c = httpx.Client(follow_redirects=True, timeout=15, headers={"User-Agent": "Mozilla/5.0"})

# 1. Check new Building Permits Socrata dataset (ydr8-5enu)
# Note: this might be on a different Socrata domain since our catalog search returned it
print("=== NEW BUILDING PERMITS DATASET ===")

# Try on the Dallas domain first  
r = c.get("https://www.dallasopendata.com/resource/ydr8-5enu.json",
          params={"$limit": "2", "$order": ":id DESC"})
print(f"dallasopendata.com: {r.status_code}")
if r.status_code == 200:
    data = r.json()
    if data:
        print(f"  Keys: {list(data[0].keys())[:15]}")
        print(f"  Sample: {data[0]}")
    else:
        print("  No data")

# Try getting metadata to find the right domain
r2 = c.get("https://api.us.socrata.com/api/catalog/v1",
           params={"ids": "ydr8-5enu"})
if r2.status_code == 200:
    results = r2.json().get("results", [])
    if results:
        domain = results[0].get("metadata", {}).get("domain", "")
        name = results[0].get("resource", {}).get("name", "")
        print(f"\n  Found: {name} on {domain}")
        # Try with the right domain
        r3 = c.get(f"https://{domain}/resource/ydr8-5enu.json",
                   params={"$limit": "2", "$order": ":id DESC"})
        print(f"  {domain}: {r3.status_code}")
        if r3.status_code == 200:
            data = r3.json()
            if data:
                print(f"  Keys: {list(data[0].keys())[:15]}")
                print(f"  Sample: {data[0]}")

# 2. Try the Dallas monthly XLSX permit table reports
print("\n=== DALLAS MONTHLY XLSX REPORTS ===")
import re
# Get the reports page
r = c.get("https://dallascityhall.com/departments/sustainabledevelopment/Pages/Reports.aspx")
if r.status_code == 200:
    # Find xlsx links
    xlsx_links = re.findall(r'href="([^"]*\.xlsx[^"]*)"', r.text, re.I)
    print(f"Found {len(xlsx_links)} xlsx links:")
    for l in xlsx_links:
        print(f"  {l}")
    
    # Find all permit report links
    all_links = re.findall(r'href="([^"]*permit_reports[^"]*)"', r.text, re.I)
    print(f"\nPermit report links: {len(all_links)}")
    for l in all_links:
        print(f"  {l}")

# 3. Try to download one of the recent xlsx files
print("\n=== DOWNLOAD TEST ===")
# Try Jan/Feb/Mar 2026
test_urls = [
    "https://dallascityhall.com/departments/sustainabledevelopment/buildinginspection/DCH%20documents/permit_reports/FY%2025-26%20Reports/January2026Permits_Table.xlsx",
    "https://dallascityhall.com/departments/sustainabledevelopment/buildinginspection/DCH%20documents/permit_reports/FY%2025-26%20Reports/February2026Permits_Table.xlsx",
    "https://dallascityhall.com/departments/sustainabledevelopment/buildinginspection/DCH%20documents/permit_reports/FY%2025-26%20Reports/March2026Permits_Table.xlsx",
    # Also try FY 24-25 format for recent months
    "https://dallascityhall.com/departments/sustainabledevelopment/buildinginspection/DCH%20documents/permit_reports/FY%2024-25%20Reports/January2025Permits_Table.xlsx",
]
for url in test_urls:
    try:
        r = c.get(url)
        print(f"  {url.split('/')[-1]}: {r.status_code} (len={len(r.content)})")
        if r.status_code == 200 and len(r.content) > 1000:
            import io
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
            ws = wb[wb.sheetnames[0]]
            print(f"    Sheet: {wb.sheetnames[0]} ({ws.max_row} rows x {ws.max_column} cols)")
            # Print header
            for row_idx in range(1, min(4, ws.max_row + 1)):
                vals = [cell.value for cell in ws[row_idx]]
                print(f"    Row {row_idx}: {vals[:12]}")
    except Exception as e:
        print(f"  {url.split('/')[-1]}: {e}")

c.close()
print("\nDone.")
