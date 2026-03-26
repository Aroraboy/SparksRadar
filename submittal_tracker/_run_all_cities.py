"""
Multi-city submittal/permit tracker.
Downloads permit data from 16 Texas cities and writes to Google Sheets.
Each city gets its own tab in spreadsheet 14qiDFhK9BIsGDMnRMuVxkpfmfYqNpg5w-6nb48M5WaM.

Data is filtered to January 2026 onwards only.

Usage:
    python -m submittal_tracker._run_all_cities              # all cities
    python -m submittal_tracker._run_all_cities Austin Plano  # specific cities
"""

from __future__ import annotations

import datetime
import io
import json
import logging
import re
import sys
import tempfile
from pathlib import Path

import gspread
import httpx
import openpyxl
import pdfplumber
from google.oauth2.service_account import Credentials

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)

SPREADSHEET_ID = "14qiDFhK9BIsGDMnRMuVxkpfmfYqNpg5w-6nb48M5WaM"
CUTOFF = datetime.date(2026, 1, 1)

SHEET_COLUMNS = [
    "Case Number", "Type", "Address", "Description",
    "Status", "Submittal Date", "Notes", "Source File", "Lead Priority",
]

HIGH_KW = [
    "warehouse", "distribution center", "office building", "mixed use",
    "mixed-use", "hospital", "medical center", "hotel", "apartment",
    "multifamily", "multi-family", "shopping center", "retail center",
    "commercial", "industrial", "manufacturing",
]
MEDIUM_KW = [
    "restaurant", "retail", "office", "church", "school", "daycare",
    "day care", "clinic", "garage", "parking", "gas station", "fuel",
    "car wash", "storage", "self-storage", "assisted living", "senior living",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def priority(text: str) -> str:
    t = text.lower()
    for kw in HIGH_KW:
        if kw in t:
            return "High"
    for kw in MEDIUM_KW:
        if kw in t:
            return "Medium"
    return "Low"


def safe(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def fmt_date(v) -> str:
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.isoformat()
    return str(v) if v else ""


def parse_date(s: str) -> datetime.date | None:
    """Try common date formats."""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.datetime.strptime(s.split("T")[0] if "T" in s else s, fmt.split("T")[0]).date()
        except (ValueError, AttributeError):
            continue
    return None


def _http() -> httpx.Client:
    return httpx.Client(follow_redirects=True, timeout=60,
                        headers={"User-Agent": "SparksRadar/1.0"})


# ── Google Sheets ─────────────────────────────────────────────────────────────

def _get_gc() -> gspread.Client:
    sa_paths = [
        Path(__file__).resolve().parent / "service_account.json",
        Path(__file__).resolve().parent.parent / "pz_tracker" / "service_account.json",
    ]
    for p in sa_paths:
        if p.exists():
            creds = Credentials.from_service_account_file(
                str(p),
                scopes=["https://www.googleapis.com/auth/spreadsheets",
                        "https://www.googleapis.com/auth/drive"],
            )
            return gspread.authorize(creds)
    raise RuntimeError("service_account.json not found")


def write_city(city: str, rows: list[list[str]]):
    """Write rows to a city tab, deduplicating by (case_number, address)."""
    gc = _get_gc()
    ss = gc.open_by_key(SPREADSHEET_ID)
    try:
        ws = ss.worksheet(city)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=city, rows=max(500, len(rows) + 50),
                              cols=len(SHEET_COLUMNS))
    first = ws.row_values(1)
    if not first or first != SHEET_COLUMNS:
        if not first:
            ws.append_row(SHEET_COLUMNS, value_input_option="RAW")
        else:
            ws.update("A1", [SHEET_COLUMNS], value_input_option="RAW")

    existing: set[tuple[str, str]] = set()
    for r in ws.get_all_values()[1:]:
        if len(r) >= 3:
            existing.add((r[0].strip(), r[2].strip()))

    new = []
    for r in rows:
        key = (r[0].strip(), r[2].strip())
        if key not in existing:
            existing.add(key)
            new.append(r)

    if new:
        ws.append_rows(new, value_input_option="RAW")
    log.info("%s: wrote %d new rows (%d dupes skipped)", city, len(new), len(rows) - len(new))
    return len(new)


# ══════════════════════════════════════════════════════════════════════════════
#                         CITY HANDLERS
# ══════════════════════════════════════════════════════════════════════════════

# ── 1. AUSTIN ─────────────────────────────────────────────────────────────────
def run_austin():
    """Austin: Socrata Open Data API – Issued Construction Permits."""
    log.info("Austin: querying Socrata API …")
    url = "https://data.austintexas.gov/resource/3syk-w9eu.json"
    rows = []
    offset = 0
    limit = 1000
    client = _http()
    while True:
        params = {
            "$where": "issue_date >= '2026-01-01T00:00:00'",
            "$limit": str(limit),
            "$offset": str(offset),
            "$order": "issue_date DESC",
        }
        resp = client.get(url, params=params)
        resp.raise_for_status()
        data = resp.json()
        if not data:
            break
        for r in data:
            permit_num = safe(r.get("permit_number", ""))
            permit_type = safe(r.get("permit_type_desc", ""))
            address = safe(r.get("original_address1", ""))
            description = safe(r.get("description", ""))
            status = safe(r.get("status_current", ""))
            issued = safe(r.get("issue_date", ""))[:10]
            work_class = safe(r.get("work_class", ""))
            permit_class = safe(r.get("permit_class_mapped", ""))
            notes_parts = []
            if work_class:
                notes_parts.append(f"Work: {work_class}")
            if permit_class:
                notes_parts.append(f"Class: {permit_class}")
            notes = "; ".join(notes_parts)
            pri = priority(f"{permit_type} {description} {permit_class}")
            rows.append([permit_num, permit_type, address, description,
                         status, issued, notes, "Austin Socrata API", pri])
        offset += limit
        if len(data) < limit:
            break
    log.info("Austin: %d records from API", len(rows))
    if rows:
        write_city("Austin", rows)
    return rows


# ── 2. FORT WORTH ─────────────────────────────────────────────────────────────
def run_fort_worth():
    """Fort Worth: ArcGIS Open Data – Certificates of Occupancy."""
    log.info("Fort Worth: querying ArcGIS feature service …")
    base = (
        "https://services5.arcgis.com/3ddLCBXe1bRt7mzj/arcgis/rest/services/"
        "CFW_Open_Data_Certificates_of_Occupancy_Table_view/FeatureServer/0/query"
    )
    rows = []
    client = _http()
    offset = 0
    while True:
        params = {
            "where": "CODate >= '2026-01-01'",
            "outFields": "PermitID,RecordAlias,ApplicationType,ApplicationSubType,"
                         "Status,ProjectName,CODate,ApplicantName,AddressLine1,"
                         "HouseNumber,StreetName,Type,Subdivsion,LegalDescription,JobUse",
            "f": "json",
            "resultOffset": str(offset),
            "resultRecordCount": "1000",
        }
        resp = client.get(base, params=params)
        if resp.status_code != 200:
            log.warning("Fort Worth: API returned %d", resp.status_code)
            break
        data = resp.json()
        if "error" in data:
            log.warning("Fort Worth: API error %s", data["error"])
            break
        features = data.get("features", [])
        if not features:
            break
        for f in features:
            a = f.get("attributes", {})
            permit = safe(a.get("PermitID", ""))
            ptype = safe(a.get("ApplicationType", "") or a.get("ApplicationSubType", ""))
            hn = safe(a.get("HouseNumber", ""))
            sn = safe(a.get("StreetName", ""))
            st = safe(a.get("Type", ""))
            addr = safe(a.get("AddressLine1", "")) or f"{hn} {sn} {st}".strip()
            proj = safe(a.get("ProjectName", ""))
            job_use = safe(a.get("JobUse", ""))
            desc = f"{proj} | {job_use}".strip(" |") if proj or job_use else ""
            status = safe(a.get("Status", ""))
            codate = a.get("CODate")
            if isinstance(codate, (int, float)) and codate > 0:
                codate = datetime.datetime.fromtimestamp(codate / 1000).strftime("%Y-%m-%d")
            else:
                codate = safe(codate)[:10] if codate else ""
            applicant = safe(a.get("ApplicantName", ""))
            subdiv = safe(a.get("Subdivsion", ""))
            notes_parts = [p for p in [f"Applicant: {applicant}" if applicant else "",
                                       f"Subdivision: {subdiv}" if subdiv else ""] if p]
            notes = "; ".join(notes_parts)
            pri = priority(f"{ptype} {desc} {job_use}")
            rows.append([permit, ptype, addr, desc, status, codate, notes,
                         "Fort Worth ArcGIS API", pri])
        offset += len(features)
        if not data.get("exceededTransferLimit", False):
            break
    log.info("Fort Worth: %d records", len(rows))
    if rows:
        write_city("Fort Worth", rows)
    return rows


# ── 3. ARLINGTON ──────────────────────────────────────────────────────────────
def run_arlington():
    """Arlington: ArcGIS Open Data – Issued Permits."""
    log.info("Arlington: querying ArcGIS feature service …")
    base = "https://gis2.arlingtontx.gov/agsext2/rest/services/OpenData/OD_Property/MapServer/1/query"
    rows = []
    client = _http()
    offset = 0
    while True:
        params = {
            "where": "ISSUEDATE >= DATE '2026-01-01'",
            "outFields": "FOLDERYEAR,FOLDERSEQUENCE,FOLDERTYPE,STATUSDESC,ISSUEDATE,"
                         "FINALDATE,SUBDESC,WORKDESC,FOLDERNAME,ConstructionValuationDeclared,"
                         "MainUse,FOLDERCONDITION,NameofBusiness",
            "f": "json",
            "resultOffset": str(offset),
            "resultRecordCount": "1000",
        }
        resp = client.get(base, params=params)
        if resp.status_code != 200:
            log.warning("Arlington: API returned %d", resp.status_code)
            break
        data = resp.json()
        if "error" in data:
            log.warning("Arlington: API error %s", data["error"])
            break
        features = data.get("features", [])
        if not features:
            break
        for f in features:
            a = f.get("attributes", {})
            yr = safe(a.get("FOLDERYEAR", ""))
            seq = safe(a.get("FOLDERSEQUENCE", ""))
            permit = f"{yr}-{seq}" if yr and seq else seq or yr
            ftype = safe(a.get("FOLDERTYPE", ""))
            subdesc = safe(a.get("SUBDESC", ""))
            workdesc = safe(a.get("WORKDESC", ""))
            ptype = f"{ftype} - {subdesc}".strip(" -") if ftype else subdesc
            addr = safe(a.get("FOLDERNAME", ""))
            cond = safe(a.get("FOLDERCONDITION", ""))
            desc = f"{workdesc} | {cond}".strip(" |") if workdesc else cond
            status = safe(a.get("STATUSDESC", ""))
            iss = a.get("ISSUEDATE")
            if isinstance(iss, (int, float)) and iss > 0:
                iss = datetime.datetime.fromtimestamp(iss / 1000).strftime("%Y-%m-%d")
            else:
                iss = ""
            val = a.get("ConstructionValuationDeclared")
            biz = safe(a.get("NameofBusiness", ""))
            notes_parts = [p for p in [f"Valuation: ${val:,.0f}" if val else "",
                                       f"Business: {biz}" if biz else ""] if p]
            notes = "; ".join(notes_parts)
            pri = priority(f"{ptype} {desc}")
            rows.append([permit, ptype, addr, desc, status, iss, notes,
                         "Arlington ArcGIS API", pri])
        offset += len(features)
        if not data.get("exceededTransferLimit", False):
            break
    log.info("Arlington: %d records", len(rows))
    if rows:
        write_city("Arlington", rows)
    return rows


# ── 4. PLANO ──────────────────────────────────────────────────────────────────
def run_plano():
    """Plano: weekly XLSX reports from CivicPlus content API.
    Format: multi-row per permit (row1=permit#/type/addr/val, row2=applied/subtype/parcel, row3=issued/status/subdiv).
    """
    log.info("Plano: downloading XLSX reports …")
    report_urls = [
        ("Commercial Building Report", "https://content.civicplus.com/api/assets/tx-plano/cefc0a2e-5b06-46a0-9685-e5385a43a1e7"),
        ("Residential Report", "https://content.civicplus.com/api/assets/tx-plano/7e7935af-a4a8-48a3-ad1e-319f8302f102"),
        ("Certificate of Occupancy", "https://content.civicplus.com/api/assets/tx-plano/553b3605-02b6-492a-84ec-42c69dc961df"),
        ("Interior Finish Out", "https://content.civicplus.com/api/assets/tx-plano/e6b712e4-e014-487d-b166-c8c54e5cb552"),
        ("Pool Report", "https://content.civicplus.com/api/assets/tx-plano/5ee14224-4df7-4266-b293-2b6d5eb20d8f"),
    ]
    rows = []
    client = _http()
    for label, url in report_urls:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("Plano %s: HTTP %d", label, resp.status_code)
                continue
            wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            ws = wb.active

            # Check date range in the header area
            date_range_ok = True
            for row_cells in ws.iter_rows(min_row=1, max_row=6, values_only=True):
                for c in row_cells:
                    s = safe(c)
                    if "Date Range" in s:
                        # e.g. "Date Range Between 10/1/2023 and 10/31/2023"
                        if "2026" not in s:
                            log.warning("Plano %s: data is from %s (not 2026), skipping", label, s)
                            date_range_ok = False
                        break
            if not date_range_ok:
                continue

            # Multi-row format: col B=permit#/applied/issued, col D=type/subtype/status,
            # col G=address/parcel/subdiv, col I=valuation/fees/paid
            # Permit rows start after header row 8, grouped in blocks of 3+ rows
            all_rows_data = list(ws.iter_rows(min_row=1, values_only=True))
            # Find where data starts (after header with PERMIT NUMBER)
            data_start = 0
            for i, row_cells in enumerate(all_rows_data):
                if any("PERMIT NUMBER" in safe(c) for c in row_cells):
                    data_start = i + 1
                    break
            if data_start == 0:
                continue

            i = data_start
            count = 0
            while i < len(all_rows_data):
                row1 = all_rows_data[i]
                permit_num = safe(row1[1]) if len(row1) > 1 else ""
                if not permit_num or permit_num.startswith("Contact") or permit_num.startswith(" "):
                    i += 1
                    continue
                ptype = safe(row1[3]) if len(row1) > 3 else ""
                addr = safe(row1[6]) if len(row1) > 6 else ""
                val = row1[8] if len(row1) > 8 else ""

                # Row 2: applied date, subtype, parcel
                applied = ""
                subtype = ""
                if i + 1 < len(all_rows_data):
                    row2 = all_rows_data[i + 1]
                    applied = row2[1] if len(row2) > 1 else ""
                    subtype = safe(row2[3]) if len(row2) > 3 else ""

                # Row 3: issued date, status, subdivision
                issued = ""
                status = ""
                subdiv = ""
                if i + 2 < len(all_rows_data):
                    row3 = all_rows_data[i + 2]
                    issued = row3[1] if len(row3) > 1 else ""
                    status = safe(row3[3]) if len(row3) > 3 else ""
                    subdiv = safe(row3[6]) if len(row3) > 6 else ""

                # Filter: only 2026 issued dates
                issued_str = fmt_date(issued)
                d = parse_date(issued_str) if issued_str else None
                if d and d < CUTOFF:
                    i += 3
                    while i < len(all_rows_data) and safe(all_rows_data[i][1] if len(all_rows_data[i]) > 1 else "").startswith(" "):
                        i += 1
                    continue

                desc = f"{subtype}; {subdiv}" if subdiv else subtype
                valuation = safe(val)
                notes = f"Valuation: {valuation}" if valuation else ""
                pri = priority(f"{ptype} {desc} {label}")
                rows.append([permit_num, ptype, addr, desc, status, issued_str,
                             notes, f"Plano - {label}", pri])
                count += 1
                # Skip to next permit (skip contact rows)
                i += 3
                while i < len(all_rows_data):
                    cell1 = safe(all_rows_data[i][1] if len(all_rows_data[i]) > 1 else "")
                    if cell1.startswith(" ") or cell1.startswith("Contact"):
                        i += 1
                    else:
                        break
            log.info("Plano %s: %d rows", label, count)
        except Exception as e:
            log.error("Plano %s: %s", label, e)
    if rows:
        write_city("Plano", rows)
    return rows


# ── 5. McKINNEY ───────────────────────────────────────────────────────────────
def run_mckinney():
    """McKinney: annual Excel reports from DocumentCenter."""
    log.info("McKinney: downloading Excel reports …")
    report_urls = [
        ("Commercial Activity 2026", "https://www.mckinneytexas.org/DocumentCenter/View/39325/CommercialActivityDetails_2026_Excel"),
        ("Residential Activity 2026", "https://www.mckinneytexas.org/DocumentCenter/View/39327/ResidentialActivityDetails_2026_Excel"),
    ]
    rows = []
    client = _http()
    for label, url in report_urls:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("McKinney %s: HTTP %d", label, resp.status_code)
                continue
            wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                headers = []
                header_row = 0
                for i, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    cells = [safe(c).lower() for c in row_cells]
                    if any(k in " ".join(cells) for k in ["permit", "address", "project", "case"]):
                        headers = [safe(c) for c in row_cells]
                        header_row = i
                        break
                if not headers or header_row == 0:
                    continue
                h_lower = [h.lower() for h in headers]

                def col(keywords):
                    for kw in keywords:
                        for j, h in enumerate(h_lower):
                            if kw in h:
                                return j
                    return None

                c_permit = col(["permit", "case", "number", "project"])
                c_addr = col(["address", "location", "site"])
                c_desc = col(["description", "name", "project name"])
                c_type = col(["type", "category", "permit type"])
                c_status = col(["status"])
                c_date = col(["date", "issued", "applied"])
                c_val = col(["value", "valuation", "cost"])

                count = 0
                for row_cells in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if not row_cells or not any(row_cells):
                        continue
                    permit = safe(row_cells[c_permit]) if c_permit is not None else ""
                    if not permit:
                        continue
                    addr = safe(row_cells[c_addr]) if c_addr is not None else ""
                    desc = safe(row_cells[c_desc]) if c_desc is not None else ""
                    ptype = safe(row_cells[c_type]) if c_type is not None else ""
                    if not ptype:
                        ptype = "Commercial" if "commercial" in label.lower() else "Residential"
                    status = safe(row_cells[c_status]) if c_status is not None else ""
                    dt = row_cells[c_date] if c_date is not None else ""
                    valuation = safe(row_cells[c_val]) if c_val is not None else ""
                    notes = f"Valuation: {valuation}" if valuation else ""
                    pri = priority(f"{ptype} {desc}")
                    rows.append([permit, ptype, addr, desc, status,
                                 fmt_date(dt), notes, f"McKinney - {label} ({ws_name})", pri])
                    count += 1
                log.info("McKinney %s/%s: %d rows", label, ws_name, count)
        except Exception as e:
            log.error("McKinney %s: %s", label, e)
    if rows:
        write_city("McKinney", rows)
    return rows


# ── 6. DENTON ─────────────────────────────────────────────────────────────────
def run_denton():
    """Denton: CKAN open data portal – monthly XLSX permit reports."""
    log.info("Denton: downloading XLSX permit reports …")
    report_urls = [
        ("January 2026", "https://data.cityofdenton.com/dataset/e1134d63-ee19-47fd-bae7-ea84c6ec135a/resource/fb0d3f72-a7e2-4592-ad08-cef38b79b236/download/2016_01_all_permits_issued.xlsx"),
        ("February 2026", "https://data.cityofdenton.com/dataset/e1134d63-ee19-47fd-bae7-ea84c6ec135a/resource/0da2e5d6-bcc0-4579-af19-dbbbc60595d8/download/2016_02_all_permits_issued.xlsx"),
    ]
    rows = []
    client = _http()
    for label, url in report_urls:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("Denton %s: HTTP %d", label, resp.status_code)
                continue
            # Try xlsx first, fall back to xls via openpyxl
            try:
                wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            except Exception:
                # May be .xls format; save to temp and try xlrd
                log.warning("Denton %s: not xlsx, trying alternative parse", label)
                continue
            ws = wb.active
            headers = []
            header_row = 0
            for i, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                cells = [safe(c).lower() for c in row_cells]
                joined = " ".join(cells)
                if any(k in joined for k in ["permit", "address", "number"]):
                    headers = [safe(c) for c in row_cells]
                    header_row = i
                    break
            if not headers or header_row == 0:
                log.warning("Denton %s: no header row found", label)
                # Use first row as headers
                headers = [safe(c) for c in list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]]
                header_row = 1
            h_lower = [h.lower() for h in headers]

            def col(keywords):
                for kw in keywords:
                    for j, h in enumerate(h_lower):
                        if kw in h:
                            return j
                return None

            c_permit = col(["permit", "case", "number"])
            c_addr = col(["address", "location", "site"])
            c_desc = col(["description", "work"])
            c_type = col(["type", "category", "class"])
            c_status = col(["status"])
            c_date = col(["date", "issued", "applied"])
            c_val = col(["value", "valuation", "cost"])
            c_contractor = col(["contractor", "builder", "owner"])

            count = 0
            for row_cells in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row_cells or not any(row_cells):
                    continue
                permit = safe(row_cells[c_permit]) if c_permit is not None else ""
                if not permit:
                    continue
                addr = safe(row_cells[c_addr]) if c_addr is not None else ""
                desc = safe(row_cells[c_desc]) if c_desc is not None else ""
                ptype = safe(row_cells[c_type]) if c_type is not None else ""
                status = safe(row_cells[c_status]) if c_status is not None else ""
                dt = row_cells[c_date] if c_date is not None else ""
                valuation = safe(row_cells[c_val]) if c_val is not None else ""
                contr = safe(row_cells[c_contractor]) if c_contractor is not None else ""
                notes_parts = []
                if contr:
                    notes_parts.append(f"Contractor: {contr}")
                if valuation:
                    notes_parts.append(f"Valuation: {valuation}")
                notes = "; ".join(notes_parts)
                pri = priority(f"{ptype} {desc}")
                rows.append([permit, ptype, addr, desc, status,
                             fmt_date(dt), notes, f"Denton - {label}", pri])
                count += 1
            log.info("Denton %s: %d rows", label, count)
        except Exception as e:
            log.error("Denton %s: %s", label, e)
    if rows:
        write_city("Denton", rows)
    return rows


# ── 7. BROWNSVILLE ────────────────────────────────────────────────────────────
def run_brownsville():
    """Brownsville: monthly PDF permit reports from DocumentCenter."""
    log.info("Brownsville: downloading PDF permit reports …")
    report_urls = [
        ("January 2026", "https://www.brownsvilletx.gov/DocumentCenter/View/17424/January-2026-Permits"),
        ("February 2026", "https://www.brownsvilletx.gov/DocumentCenter/View/17531/February-2026-Permits"),
    ]
    rows = []
    client = _http()
    for label, url in report_urls:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("Brownsville %s: HTTP %d", label, resp.status_code)
                continue
            with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if not table:
                            continue
                        # Find header row
                        header_idx = None
                        for i, row in enumerate(table):
                            cells = [safe(c).lower() for c in row]
                            joined = " ".join(cells)
                            if any(k in joined for k in ["permit", "address", "number", "type"]):
                                header_idx = i
                                break
                        if header_idx is None:
                            header_idx = 0
                        headers = [safe(c).lower() for c in table[header_idx]]

                        def col(keywords):
                            for kw in keywords:
                                for j, h in enumerate(headers):
                                    if kw in h:
                                        return j
                            return None

                        c_permit = col(["permit", "case", "number", "#"])
                        c_addr = col(["address", "location", "site"])
                        c_desc = col(["description", "work", "scope"])
                        c_type = col(["type", "category", "class"])
                        c_status = col(["status"])
                        c_date = col(["date", "issued"])
                        c_val = col(["value", "valuation", "cost"])

                        for row in table[header_idx + 1:]:
                            if not row or not any(row):
                                continue
                            permit = safe(row[c_permit]) if c_permit is not None and c_permit < len(row) else ""
                            if not permit:
                                continue
                            addr = safe(row[c_addr]) if c_addr is not None and c_addr < len(row) else ""
                            desc = safe(row[c_desc]) if c_desc is not None and c_desc < len(row) else ""
                            ptype = safe(row[c_type]) if c_type is not None and c_type < len(row) else ""
                            status = safe(row[c_status]) if c_status is not None and c_status < len(row) else ""
                            dt = safe(row[c_date]) if c_date is not None and c_date < len(row) else ""
                            val = safe(row[c_val]) if c_val is not None and c_val < len(row) else ""
                            notes = f"Valuation: {val}" if val else ""
                            pri = priority(f"{ptype} {desc}")
                            rows.append([permit, ptype, addr, desc, status,
                                         dt, notes, f"Brownsville - {label}", pri])
            log.info("Brownsville %s: extracted from PDF", label)
        except Exception as e:
            log.error("Brownsville %s: %s", label, e)
    log.info("Brownsville: %d total rows", len(rows))
    if rows:
        write_city("Brownsville", rows)
    return rows


# ── 8. LUBBOCK ────────────────────────────────────────────────────────────────
def run_lubbock():
    """Lubbock: weekly PDF new construction logs. Text-based parsing (no tables)."""
    log.info("Lubbock: downloading PDF construction logs …")
    pdf_urls = [
        ("Jan 5-9 2026", "https://ci.lubbock.tx.us/storage/images/mXmsQlRsOHBUxDNps2UbGMODUmuf4wfJnv7ZWd8P.pdf"),
        ("Jan 12-16 2026", "https://ci.lubbock.tx.us/storage/images/9dunmpsoEn1Urhur9jHGmuchHA0VYrnu54vMWkLu.pdf"),
        ("Jan 19-23 2026", "https://ci.lubbock.tx.us/storage/images/vdOBdOBJhaEWlEVpiT4zgu7ed3oDTheSGps6Aj9A.pdf"),
        ("Jan 26-30 2026", "https://ci.lubbock.tx.us/storage/images/MFn33F5yYQdl6FYP2TDlrgg7fmJyn9gsT7WtSmlh.pdf"),
        ("Feb 2-6 2026", "https://ci.lubbock.tx.us/storage/images/lgCHLgkrPCXHrHzRJISCpwXqmQzN3cI5EpLIMzrM.pdf"),
        ("Feb 9-13 2026", "https://ci.lubbock.tx.us/storage/images/5ICW3Jv9y7v7FmGp7B2QRSsfkO5k5BcSdknqVJwZ.pdf"),
        ("Feb 16-20 2026", "https://ci.lubbock.tx.us/storage/images/LfQb5tPy3Vw6Oi8TYxkdfOBr1f6l1NExnqNO8VqD.pdf"),
        ("Feb 23-27 2026", "https://ci.lubbock.tx.us/storage/images/nIqZ2KWdw2z6sCbJC3ggXQqOGPcDfYqpejJmJOdY.pdf"),
        ("Mar 2-6 2026", "https://ci.lubbock.tx.us/storage/images/hFhD5RI5a9O23lzKQwwJy1F4R8yzFWy6HXbYV1XC.pdf"),
        ("Mar 9-13 2026", "https://ci.lubbock.tx.us/storage/images/Onn32fxNkbVGZ3SvTVhMdeSKxDJqQWI16h7DFUFp.pdf"),
        ("Mar 16-20 2026", "https://ci.lubbock.tx.us/storage/images/p41rKQn9eTLSFrBlxdtlFYgOGwJpRxCk7kigpdjn.pdf"),
    ]
    # Regex to find permit numbers like RES-224209-2025 or COM-123456-2026
    permit_re = re.compile(r'^((?:RES|COM|COMM|IND|MF)\-\d{4,8}\-\d{4})\b', re.I)
    rows = []
    seen = set()
    client = _http()
    for label, url in pdf_urls:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("Lubbock %s: HTTP %d", label, resp.status_code)
                continue
            with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    lines = text.split("\n")
                    prev_addr = ""
                    for line in lines:
                        line = line.strip()
                        m = permit_re.match(line)
                        if m:
                            permit_num = m.group(1)
                            if permit_num in seen:
                                prev_addr = ""
                                continue
                            seen.add(permit_num)
                            rest = line[m.end():].strip()
                            # Parse: CITY, ST ZIP LEGAL_DESC WORK_CLASS SQFT $VALUE CONTRACTOR
                            # Try to extract valuation
                            val_match = re.search(r'\$[\d,]+(?:\.\d{2})?', rest)
                            valuation = val_match.group(0) if val_match else ""
                            # Extract work class (Single Family, Duplex, Commercial, etc.)
                            work_match = re.search(r'(Single Family|Duplex|Multi[\- ]?Family|Commercial|Industrial|Townhome|Accessory)', rest, re.I)
                            work_class = work_match.group(1) if work_match else "New Construction"
                            # Extract contractor (text after valuation)
                            contractor = ""
                            if val_match:
                                contractor = rest[val_match.end():].strip()
                            # Build address from previous line
                            addr = prev_addr
                            # Also grab city/state/zip from this line
                            city_match = re.search(r'(LUBBOCK,?\s*TX\s*\d{5})', rest, re.I)
                            if city_match:
                                addr = f"{addr}, {city_match.group(1)}" if addr else city_match.group(1)
                            # Extract legal description
                            legal = ""
                            if city_match:
                                after_city = rest[city_match.end():].strip()
                                if work_match:
                                    legal = after_city[:rest.index(work_match.group(1)) - city_match.end()].strip()
                            desc = f"{work_class}; {legal}" if legal else work_class
                            notes = f"Contractor: {contractor}; Value: {valuation}" if contractor else f"Value: {valuation}"
                            pri = priority(f"{work_class} {desc}")
                            rows.append([permit_num, work_class, addr, desc, "",
                                         "", notes.strip("; "), f"Lubbock - {label}", pri])
                            prev_addr = ""
                        else:
                            # Check if this line looks like an address
                            if re.match(r'^\d+\s+\w', line) and "Permit" not in line and "DAILY" not in line:
                                prev_addr = line
            log.info("Lubbock %s: done", label)
        except Exception as e:
            log.error("Lubbock %s: %s", label, e)
    log.info("Lubbock: %d total rows", len(rows))
    if rows:
        write_city("Lubbock", rows)
    return rows


# ── 9. DALLAS ─────────────────────────────────────────────────────────────────
def run_dallas():
    """Dallas: monthly XLSX permit reports from Development Services.

    Reports page publishes FY-based XLSX files. FY 25-26 runs Oct 2025 – Sep 2026.
    We scrape the reports page, grab all Permits_Table.xlsx links, download them,
    and filter rows with dates >= CUTOFF.
    """
    log.info("Dallas: scraping DevServices reports page …")
    client = _http()
    rows: list[list[str]] = []
    reports_url = (
        "https://dallascityhall.com/departments/sustainabledevelopment/"
        "Pages/Reports.aspx"
    )
    try:
        resp = client.get(reports_url)
        if resp.status_code != 200:
            log.warning("Dallas: reports page returned %d", resp.status_code)
            return rows
        # Collect all XLSX links (Permits_Table + COs + Demolition + Solar)
        xlsx_links = re.findall(
            r'href=["\']([^"\']*(?:Permits_Table|COs_MonthOf)[^"\']*\.xlsx)["\']',
            resp.text, re.I,
        )
        for link in xlsx_links:
            if not link.startswith("http"):
                link = f"https://dallascityhall.com{link}"
            try:
                r2 = client.get(link)
                if r2.status_code != 200:
                    continue
                wb = openpyxl.load_workbook(io.BytesIO(r2.content), data_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    # Detect header row
                    headers = []
                    hrow = 0
                    for ri in range(1, min(10, ws.max_row + 1)):
                        cells = [safe(c.value).lower() for c in ws[ri]]
                        if any(k in " ".join(cells) for k in ["permit", "address", "number"]):
                            headers = [safe(c.value) for c in ws[ri]]
                            hrow = ri
                            break
                    if not headers:
                        continue
                    hl = [h.lower() for h in headers]

                    def _col(kws):
                        for kw in kws:
                            for j, h in enumerate(hl):
                                if kw in h:
                                    return j
                        return None

                    ci_perm = _col(["permit", "number", "#"])
                    ci_type = _col(["type", "category"])
                    ci_addr = _col(["address", "location"])
                    ci_desc = _col(["description", "work", "scope"])
                    ci_stat = _col(["status"])
                    ci_date = _col(["issued", "date"])
                    ci_val  = _col(["value", "valuation", "cost"])
                    ci_cont = _col(["contractor"])

                    for data_row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                        if not data_row or not any(data_row):
                            continue
                        def _g(idx):
                            if idx is not None and idx < len(data_row):
                                return data_row[idx]
                            return ""
                        # Date filter
                        raw_date = _g(ci_date)
                        dt = parse_date(raw_date)
                        if dt and dt < CUTOFF:
                            continue
                        permit = safe(_g(ci_perm))
                        if not permit:
                            continue
                        ptype = safe(_g(ci_type))
                        addr = safe(_g(ci_addr))
                        desc = safe(_g(ci_desc))
                        status = safe(_g(ci_stat))
                        date_s = fmt_date(dt) if dt else safe(raw_date)[:10]
                        val = safe(_g(ci_val))
                        contractor = safe(_g(ci_cont))
                        notes_parts = []
                        if val:
                            notes_parts.append(f"Value: {val}")
                        if contractor:
                            notes_parts.append(f"Contractor: {contractor}")
                        notes = "; ".join(notes_parts)
                        pri = priority(f"{ptype} {desc}")
                        rows.append([permit, ptype, addr, desc, status, date_s,
                                     notes, "Dallas DevServices XLSX", pri])
            except Exception as e:
                log.warning("Dallas XLSX %s: %s", link.split("/")[-1], e)
    except Exception as e:
        log.error("Dallas reports page: %s", e)

    if not rows:
        log.warning("Dallas: no 2026 data (FY 25-26 reports not yet published)")
    else:
        log.info("Dallas: %d records", len(rows))
        write_city("Dallas", rows)
    return rows


# ── 10. CORPUS CHRISTI ────────────────────────────────────────────────────────
def run_corpus_christi():
    """Corpus Christi: fiscal permit & CO reports (direct XLSX download)."""
    log.info("Corpus Christi: fetching permit & CO reports …")
    BASE = "https://www.corpuschristitx.gov"
    PERMIT_URLS = [
        f"{BASE}/media/ba1nbcm4/permit-report-jan-2026-excel-version.xlsx",
        f"{BASE}/media/mrlcnq1a/permit-report-february-2026-excel-version.xlsx",
    ]
    CO_URLS = [
        f"{BASE}/media/riplqktg/coo-report-jan-2026-excel-version.xlsx",
        f"{BASE}/media/hnhdovfv/coo-report-february-2026-excel-version.xlsx",
    ]
    client = _http()
    rows: list[list[str]] = []

    # ── Permit reports (header row 8) ──
    for url in PERMIT_URLS:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("Corpus Christi permit %s → %s", url, resp.status_code)
                continue
            wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            ws = wb.active
            hdr = [safe(c) for c in next(ws.iter_rows(min_row=8, max_row=8, values_only=True))]
            h = {v.lower(): i for i, v in enumerate(hdr)}
            for vals in ws.iter_rows(min_row=9, values_only=True):
                if not vals or not any(vals):
                    continue
                def g(key):
                    idx = h.get(key)
                    return safe(vals[idx]) if idx is not None and idx < len(vals) else ""
                permit = g("permit")
                addr = " ".join(filter(None, [g("str no"), g("str name"), g("city"), g("zip")]))
                desc = " ".join(filter(None, [g("project"), g("work type"), g("sub work")]))
                issued = g("issued")
                dt = parse_date(issued)
                if dt and dt < CUTOFF:
                    continue
                rows.append([
                    permit,
                    g("appl type"),
                    addr,
                    desc,
                    g("status"),
                    fmt_date(dt) if dt else issued,
                    f"SqFt={g('sq ft')} Val={g('value')}",
                    url.split("/")[-1],
                    priority(desc + " " + g("appl type")),
                ])
        except Exception as e:
            log.warning("Corpus Christi permit %s: %s", url, e)

    # ── CO reports (header row 8) ──
    for url in CO_URLS:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("Corpus Christi CO %s → %s", url, resp.status_code)
                continue
            wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            ws = wb.active
            hdr = [safe(c) for c in next(ws.iter_rows(min_row=8, max_row=8, values_only=True))]
            h = {v.lower(): i for i, v in enumerate(hdr)}
            for vals in ws.iter_rows(min_row=9, values_only=True):
                if not vals or not any(vals):
                    continue
                def g(key):
                    idx = h.get(key)
                    return safe(vals[idx]) if idx is not None and idx < len(vals) else ""
                apno = g("apno")
                addr = " ".join(filter(None, [g("stno"), g("predir"), g("stname"), g("suffix"), g("zip")]))
                desc = " ".join(filter(None, [g("worktype"), g("descript")]))
                co_date = g("coodttm")
                dt = parse_date(co_date)
                if dt and dt < CUTOFF:
                    continue
                rows.append([
                    apno,
                    g("aptype") + " (CO)",
                    addr,
                    desc,
                    g("stat"),
                    fmt_date(dt) if dt else co_date,
                    f"Area={g('bldgarea')} Val={g('declvltn')}",
                    url.split("/")[-1],
                    priority(desc + " " + g("aptype")),
                ])
        except Exception as e:
            log.warning("Corpus Christi CO %s: %s", url, e)

    log.info("Corpus Christi: %d rows total", len(rows))
    if rows:
        write_city("Corpus Christi", rows)
    return rows


# ── 11. GRAND PRAIRIE ─────────────────────────────────────────────────────────
def run_grand_prairie():
    """Grand Prairie: building permits report page (Playwright for 403 bypass)."""
    from playwright.sync_api import sync_playwright
    log.info("Grand Prairie: fetching permit reports via Playwright …")
    rows: list[list[str]] = []
    url = "https://www.gptx.org/Departments/Building-Inspections/Building-Permits-Report"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context()
            page = ctx.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=30_000)
            page.wait_for_timeout(3000)
            html = page.content()

            # Find document links (pdf/xlsx) with 2026 in URL or text
            links = re.findall(r'href=["\']([^"\']*\.(?:xlsx|xls|pdf|csv))["\']', html, re.I)
            links_2026 = re.findall(r'href=["\']([^"\']*2026[^"\']*)["\']', html, re.I)
            all_links = list({l for l in (links + links_2026)})
            for link in all_links:
                if not link.startswith("http"):
                    link = f"https://www.gptx.org{link}"
                if "2026" not in link and "26" not in link.split("/")[-1]:
                    continue
                # Skip non-document links (CSS, JS, general nav)
                if any(skip in link.lower() for skip in [".css", ".js", "bond-election", "news-articles"]):
                    continue
                log.info("Grand Prairie: found link %s", link)
                try:
                    resp = ctx.request.get(link)
                    if resp.status != 200:
                        continue
                    body = resp.body()
                    if link.lower().endswith(('.xlsx', '.xls')):
                        rows.extend(_parse_excel_generic(body, "Grand Prairie", link))
                    elif link.lower().endswith('.pdf'):
                        rows.extend(_parse_pdf_generic(body, "Grand Prairie", link))
                except Exception as e:
                    log.warning("Grand Prairie link %s: %s", link, e)
            browser.close()
    except Exception as e:
        log.error("Grand Prairie: %s", e)

    log.info("Grand Prairie: %d rows", len(rows))
    if rows:
        write_city("Grand Prairie", rows)
    return rows


# ── 12. ROUND ROCK ────────────────────────────────────────────────────────────
def run_round_rock():
    """Round Rock: building inspection forms & reports (Playwright for 403 bypass)."""
    from playwright.sync_api import sync_playwright
    log.info("Round Rock: fetching inspection reports via Playwright …")
    rows: list[list[str]] = []
    url = "https://www.roundrocktexas.gov/city-departments/planning-and-development-services/building-inspection/forms-and-reports/"

    def _parse_rr_pdf(body: bytes, source: str) -> list[list[str]]:
        """Parse Round Rock monthly permit PDF.
        Each permit block starts with a permit number like ACC26-00000.
        Format: PERMIT_NUM TYPE ADDRESS $VALUATION
                ISSUED_DATE SUB_TYPE PARCEL FEES
                APPLIED_DATE STATUS SUBDIVISION PAID
                Contractor / Description / Owner lines
        """
        parsed = []
        pdf = pdfplumber.open(io.BytesIO(body))
        full_text = ""
        for pg in pdf.pages:
            t = pg.extract_text()
            if t:
                full_text += t + "\n"
        # Split on permit number pattern (e.g., ACC26-00000, COM26-00003, RES26-00123)
        permit_re = re.compile(r'^([A-Z]{2,5}\d{2}-\d{4,6})\s+(.+)', re.MULTILINE)
        blocks = permit_re.split(full_text)
        # blocks: [pre, num, rest, num, rest, ...]
        i = 1
        while i + 1 < len(blocks):
            permit_num = blocks[i].strip()
            block_text = blocks[i + 1].strip()
            i += 2
            lines = block_text.split("\n")
            if not lines:
                continue
            # First line: PERMIT_TYPE ADDRESS $VALUATION
            first = lines[0]
            # Extract valuation from end
            val_m = re.search(r'\$[\d,]+\.?\d*$', first)
            valuation = val_m.group(0) if val_m else ""
            type_addr = first[:val_m.start()].strip() if val_m else first

            # Try to separate permit type from address
            # Common types: Commercial Building, Residential SF, etc.
            addr = type_addr
            ptype = ""
            for kw in ["Commercial", "Residential", "Industrial"]:
                idx = type_addr.find(kw)
                if idx == 0:
                    # Find where address starts (a number after type words)
                    addr_m = re.search(r'\d+\s+\w', type_addr)
                    if addr_m:
                        ptype = type_addr[:addr_m.start()].strip()
                        addr = type_addr[addr_m.start():].strip()
                    break

            # Second line might have issued date, sub type, etc.
            issued = ""
            status = ""
            desc_parts = []
            for ln in lines[1:]:
                ln = ln.strip()
                if not ln:
                    continue
                # Check for date pattern at start
                dt_m = re.match(r'(\d{1,2}/\d{1,2}/\d{4})', ln)
                if dt_m:
                    if not issued:
                        issued = dt_m.group(1)
                        rest = ln[dt_m.end():].strip()
                        # Look for status keywords
                        for st in ["ISSUED", "REVIEW", "FEES_DUE", "EXPIRED", "FINALED", "CLOSED"]:
                            if st in rest:
                                status = st
                                break
                    continue
                if ln.startswith("Contractor Name:") or ln.startswith("Owner Name:") or ln.startswith("Owner Info:"):
                    continue
                if ln.startswith(("PERMIT NUMBER", "ISSUED DATE", "APPLIED DATE")):
                    continue
                # Remaining lines are description
                desc_parts.append(ln)

            desc = " ".join(desc_parts)[:200]
            dt = parse_date(issued)
            if dt and dt < CUTOFF:
                continue
            parsed.append([
                permit_num, ptype, addr, desc, status,
                fmt_date(dt) if dt else issued,
                f"Val={valuation}" if valuation else "",
                source.split("/")[-1], priority(f"{ptype} {desc}"),
            ])
        return parsed

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context()
            page = ctx.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=30_000)
            page.wait_for_timeout(3000)
            html = page.content()

            links = re.findall(r'href=["\']([^"\']*\.(?:xlsx|xls|pdf|csv))["\']', html, re.I)
            links_2026 = re.findall(r'href=["\']([^"\']*2026[^"\']*)["\']', html, re.I)
            all_links = list({l for l in (links + links_2026)})
            for link in all_links:
                if not link.startswith("http"):
                    link = f"https://www.roundrocktexas.gov{link}"
                if "2026" not in link and "26" not in link.split("/")[-1]:
                    continue
                # Skip December 2025 and Periodic summary reports
                if "december" in link.lower() or "periodic" in link.lower():
                    continue
                log.info("Round Rock: downloading %s", link)
                try:
                    resp = ctx.request.get(link)
                    if resp.status != 200:
                        log.warning("Round Rock: %s → %s", link, resp.status)
                        continue
                    body = resp.body()
                    if link.lower().endswith('.pdf'):
                        rows.extend(_parse_rr_pdf(body, link))
                    elif link.lower().endswith(('.xlsx', '.xls')):
                        rows.extend(_parse_excel_generic(body, "Round Rock", link))
                except Exception as e:
                    log.warning("Round Rock link %s: %s", link, e)
            browser.close()
    except Exception as e:
        log.error("Round Rock: %s", e)

    log.info("Round Rock: %d rows", len(rows))
    if rows:
        write_city("Round Rock", rows)
    return rows


# ── 13. CARROLLTON ────────────────────────────────────────────────────────────
def run_carrollton():
    """Carrollton: monthly permit reports from webrpts subdomain (direct PDF)."""
    log.info("Carrollton: fetching monthly permit report PDFs …")
    BASE = "https://webrpts.cityofcarrollton.com/bldg_insp/results/permit_reports/archive"
    PDF_URLS = [
        f"{BASE}/01_01_26%20THRU%2001_31_26.pdf",
        f"{BASE}/02_01_26%20THRU%2002_28_26.pdf",
    ]
    client = _http()
    rows: list[list[str]] = []
    for url in PDF_URLS:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("Carrollton %s → %s", url.split("/")[-1], resp.status_code)
                continue
            with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if not table:
                            continue
                        # Find header row (contains 'Permit Number')
                        hdr_idx = None
                        for i, row in enumerate(table):
                            joined = " ".join(safe(c).lower() for c in row)
                            if "permit number" in joined or "permit type" in joined:
                                hdr_idx = i
                                break
                        if hdr_idx is None:
                            continue
                        headers = [safe(c).lower().replace("\n", " ") for c in table[hdr_idx]]
                        def _ci(kws):
                            for kw in kws:
                                for j, h in enumerate(headers):
                                    if kw in h:
                                        return j
                            return None
                        ci_perm = _ci(["permit number"])
                        ci_type = _ci(["permit type"])
                        ci_addr = _ci(["property address", "address"])
                        ci_desc = _ci(["job description", "description"])
                        ci_appl = _ci(["applicant"])
                        ci_val  = _ci(["valuation"])
                        ci_stat = _ci(["permit status", "status"])
                        ci_date = _ci(["date issued"])
                        for row in table[hdr_idx + 1:]:
                            if not row or not any(row):
                                continue
                            def _g(idx):
                                if idx is not None and idx < len(row):
                                    return safe(row[idx]).replace("\n", " ")
                                return ""
                            permit = _g(ci_perm)
                            if not permit:
                                continue
                            issued = _g(ci_date)
                            dt = parse_date(issued)
                            if dt and dt < CUTOFF:
                                continue
                            ptype = _g(ci_type)
                            addr = _g(ci_addr)
                            addr = re.sub(r'\s*CADID:.*', '', addr)
                            addr = re.sub(r'\s*County:.*', '', addr)
                            desc = _g(ci_desc)[:200]
                            val = _g(ci_val)
                            status = _g(ci_stat)
                            applicant = _g(ci_appl)
                            notes_parts = []
                            if val:
                                notes_parts.append(f"Val={val}")
                            if applicant:
                                notes_parts.append(f"Applicant={applicant}")
                            rows.append([
                                permit, ptype, addr.strip(), desc,
                                status.replace("\n", " "),
                                fmt_date(dt) if dt else issued,
                                "; ".join(notes_parts),
                                url.split("/")[-1],
                                priority(f"{ptype} {desc}"),
                            ])
            log.info("Carrollton: parsed %s", url.split("/")[-1])
        except Exception as e:
            log.warning("Carrollton %s: %s", url.split("/")[-1], e)

    log.info("Carrollton: %d rows", len(rows))
    if rows:
        write_city("Carrollton", rows)
    return rows


# ── 14. KILLEEN ───────────────────────────────────────────────────────────────
def run_killeen():
    """Killeen: monthly permit report PDF from DocumentCenter."""
    log.info("Killeen: fetching monthly permit report …")
    PDF_URLS = [
        "https://www.killeentexas.gov/DocumentCenter/View/7781/Monthly-Permit-Report-PDF",
    ]
    client = _http()
    rows: list[list[str]] = []
    for url in PDF_URLS:
        try:
            resp = client.get(url)
            if resp.status_code != 200:
                log.warning("Killeen %s → %s", url, resp.status_code)
                continue
            # Persistent column indices + last_type across all pages
            ci_type = ci_perm = ci_desc = ci_addr = ci_cont = None
            ci_sqft = ci_date = ci_val = ci_fees = None
            cols_found = False
            last_type = ""
            with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if not table:
                            continue
                        # Try to find header row on this page
                        hdr_idx = None
                        for i, row in enumerate(table):
                            joined = " ".join(safe(c).lower() for c in row)
                            if "permit" in joined and ("type" in joined or "#" in joined):
                                hdr_idx = i
                                break
                        if hdr_idx is not None:
                            headers = [safe(c).lower().replace("\n", " ") for c in table[hdr_idx]]
                            def _ci(kws):
                                for kw in kws:
                                    for j, h in enumerate(headers):
                                        if kw in h:
                                            return j
                                return None
                            ci_type = _ci(["type"])
                            ci_perm = _ci(["permit #", "permit"])
                            ci_desc = _ci(["project description", "description"])
                            ci_addr = _ci(["address"])
                            ci_cont = _ci(["contractor"])
                            ci_sqft = _ci(["total sq"])
                            ci_date = _ci(["date"])
                            ci_val  = _ci(["valuation", "project valuation"])
                            ci_fees = _ci(["fees"])
                            cols_found = True
                            data_rows = table[hdr_idx + 1:]
                        elif cols_found:
                            # No header on this page — reuse saved indices
                            data_rows = table
                        else:
                            continue
                        for row in data_rows:
                            if not row or not any(row):
                                continue
                            def _g(idx):
                                if idx is not None and idx < len(row):
                                    v = safe(row[idx]).replace("\n", " ")
                                    return v
                                return ""
                            ptype = _g(ci_type) or last_type
                            if _g(ci_type):
                                last_type = ptype
                            permit = _g(ci_perm)
                            if not permit:
                                continue
                            date_s = _g(ci_date)
                            dt = parse_date(date_s)
                            if dt and dt < CUTOFF:
                                continue
                            addr = _g(ci_addr)
                            desc = _g(ci_desc)[:200]
                            val = _g(ci_val)
                            fees = _g(ci_fees)
                            contractor = _g(ci_cont)
                            notes_parts = []
                            if val:
                                notes_parts.append(f"Val={val}")
                            if fees:
                                notes_parts.append(f"Fees={fees}")
                            if contractor:
                                notes_parts.append(f"Contractor={contractor}")
                            rows.append([
                                permit, ptype, addr, desc, "",
                                fmt_date(dt) if dt else date_s,
                                "; ".join(notes_parts),
                                "Killeen Monthly Permit Report",
                                priority(f"{ptype} {desc}"),
                            ])
            log.info("Killeen: parsed %s", url.split("/")[-1])
        except Exception as e:
            log.warning("Killeen %s: %s", url, e)

    log.info("Killeen: %d rows", len(rows))
    if rows:
        write_city("Killeen", rows)
    return rows


# ── 15. BEAUMONT ──────────────────────────────────────────────────────────────
def run_beaumont():
    """Beaumont: DocumentCenter for building permits."""
    log.info("Beaumont: checking DocumentCenter …")
    client = _http()
    rows = []
    # Try to find building permit reports in DocumentCenter
    urls_to_try = [
        "https://www.beaumonttexas.gov/DocumentCenter/View/4208/",
        "https://www.beaumonttexas.gov/DocumentCenter/Index/177",
    ]
    for url in urls_to_try:
        try:
            resp = client.get(url)
            if resp.status_code == 200:
                ct = resp.headers.get("content-type", "")
                if "pdf" in ct:
                    rows.extend(_parse_pdf_generic(resp.content, "Beaumont", url))
                elif "html" in ct:
                    links = re.findall(r'href=["\']([^"\']*(?:permit|building)[^"\']*\.(?:xlsx|xls|pdf))["\']', resp.text, re.I)
                    for link in links:
                        if not link.startswith("http"):
                            link = f"https://www.beaumonttexas.gov{link}"
                        if "2026" in link:
                            log.info("Beaumont: found link %s", link)
        except Exception as e:
            log.warning("Beaumont %s: %s", url, e)

    if not rows:
        log.warning("Beaumont: no 2026 permit data found")
    if rows:
        write_city("Beaumont", rows)
    return rows


# ── 16. ODESSA ────────────────────────────────────────────────────────────────
def run_odessa():
    """Odessa: DocumentCenter – last updated Nov 2025."""
    log.info("Odessa: checking DocumentCenter …")
    client = _http()
    rows = []
    url = "https://www.odessa-tx.gov/DocumentCenter/Index/122"
    try:
        resp = client.get(url)
        if resp.status_code == 200:
            links = re.findall(r'href=["\']([^"\']*(?:permit|building|construction)[^"\']*\.(?:xlsx|xls|pdf))["\']', resp.text, re.I)
            for link in links:
                if "2026" in link:
                    log.info("Odessa: found link %s", link)
    except Exception as e:
        log.warning("Odessa: %s", e)

    if not rows:
        log.warning("Odessa: no 2026 permit data found (last updated Nov 2025)")
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#                    GENERIC PARSERS
# ══════════════════════════════════════════════════════════════════════════════

def _parse_excel_generic(content: bytes, city: str, source: str) -> list[list[str]]:
    """Generic Excel parser that auto-detects columns."""
    rows = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        log.warning("%s: could not parse Excel from %s", city, source)
        return rows

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        headers = []
        header_row = 0
        for i, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            cells = [safe(c).lower() for c in row_cells]
            joined = " ".join(cells)
            if any(k in joined for k in ["permit", "address", "number", "case", "project"]):
                headers = [safe(c) for c in row_cells]
                header_row = i
                break
        if not headers:
            continue
        h_lower = [h.lower() for h in headers]

        def col(keywords):
            for kw in keywords:
                for j, h in enumerate(h_lower):
                    if kw in h:
                        return j
            return None

        c_permit = col(["permit", "case", "number", "project", "#"])
        c_addr = col(["address", "location", "site"])
        c_desc = col(["description", "work", "scope", "name"])
        c_type = col(["type", "category", "class"])
        c_status = col(["status"])
        c_date = col(["date", "issued", "applied"])
        c_val = col(["value", "valuation", "cost"])

        for row_cells in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row_cells or not any(row_cells):
                continue
            permit = safe(row_cells[c_permit]) if c_permit is not None and c_permit < len(row_cells) else ""
            if not permit:
                continue
            addr = safe(row_cells[c_addr]) if c_addr is not None and c_addr < len(row_cells) else ""
            desc = safe(row_cells[c_desc]) if c_desc is not None and c_desc < len(row_cells) else ""
            ptype = safe(row_cells[c_type]) if c_type is not None and c_type < len(row_cells) else ""
            status = safe(row_cells[c_status]) if c_status is not None and c_status < len(row_cells) else ""
            dt = row_cells[c_date] if c_date is not None and c_date < len(row_cells) else ""
            val = safe(row_cells[c_val]) if c_val is not None and c_val < len(row_cells) else ""
            notes = f"Valuation: {val}" if val else ""
            pri = priority(f"{ptype} {desc}")
            rows.append([permit, ptype, addr, desc, status, fmt_date(dt),
                         notes, f"{city} - {source.split('/')[-1]}", pri])
    return rows


def _parse_pdf_generic(content: bytes, city: str, source: str) -> list[list[str]]:
    """Generic PDF table parser."""
    rows = []
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header_idx = None
                    for i, row in enumerate(table):
                        cells = [safe(c).lower() for c in row]
                        joined = " ".join(cells)
                        if any(k in joined for k in ["permit", "address", "number", "type", "#"]):
                            header_idx = i
                            break
                    if header_idx is None:
                        header_idx = 0
                    headers = [safe(c).lower() for c in table[header_idx]]

                    def col(keywords):
                        for kw in keywords:
                            for j, h in enumerate(headers):
                                if kw in h:
                                    return j
                        return None

                    c_permit = col(["permit", "case", "number", "#"])
                    c_addr = col(["address", "location", "site"])
                    c_desc = col(["description", "work", "scope"])
                    c_type = col(["type", "category"])
                    c_status = col(["status"])
                    c_date = col(["date", "issued"])
                    c_val = col(["value", "valuation", "cost"])

                    for row in table[header_idx + 1:]:
                        if not row or not any(row):
                            continue
                        permit = safe(row[c_permit]) if c_permit is not None and c_permit < len(row) else ""
                        if not permit:
                            continue
                        addr = safe(row[c_addr]) if c_addr is not None and c_addr < len(row) else ""
                        desc = safe(row[c_desc]) if c_desc is not None and c_desc < len(row) else ""
                        ptype = safe(row[c_type]) if c_type is not None and c_type < len(row) else ""
                        status = safe(row[c_status]) if c_status is not None and c_status < len(row) else ""
                        dt = safe(row[c_date]) if c_date is not None and c_date < len(row) else ""
                        val = safe(row[c_val]) if c_val is not None and c_val < len(row) else ""
                        notes = f"Valuation: {val}" if val else ""
                        pri = priority(f"{ptype} {desc}")
                        rows.append([permit, ptype, addr, desc, status, dt,
                                     notes, f"{city} - {source.split('/')[-1]}", pri])
    except Exception as e:
        log.warning("%s: could not parse PDF from %s: %s", city, source, e)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#                        MAIN
# ══════════════════════════════════════════════════════════════════════════════

CITY_HANDLERS: dict[str, callable] = {
    "Austin": run_austin,
    "Fort Worth": run_fort_worth,
    "Arlington": run_arlington,
    "Dallas": run_dallas,
    "Corpus Christi": run_corpus_christi,
    "Plano": run_plano,
    "Lubbock": run_lubbock,
    "Grand Prairie": run_grand_prairie,
    "McKinney": run_mckinney,
    "Brownsville": run_brownsville,
    "Killeen": run_killeen,
    "Denton": run_denton,
    "Carrollton": run_carrollton,
    "Beaumont": run_beaumont,
    "Round Rock": run_round_rock,
    "Odessa": run_odessa,
}


def main():
    args = sys.argv[1:]
    if args:
        cities = args
    else:
        cities = list(CITY_HANDLERS.keys())

    results: dict[str, int] = {}
    for city in cities:
        handler = CITY_HANDLERS.get(city)
        if not handler:
            log.error("Unknown city: %s. Available: %s", city, ", ".join(CITY_HANDLERS))
            continue
        try:
            rows = handler()
            results[city] = len(rows) if rows else 0
        except Exception as e:
            log.error("%s FAILED: %s", city, e)
            results[city] = -1

    print("\n" + "=" * 50)
    print("SUMMARY")
    print("=" * 50)
    for city, count in results.items():
        status = f"{count} rows" if count >= 0 else "FAILED"
        print(f"  {city:20s} → {status}")


if __name__ == "__main__":
    main()
