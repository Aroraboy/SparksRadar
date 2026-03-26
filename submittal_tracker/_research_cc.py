"""Inspect CC Excel permit reports + test Playwright for 403 cities."""
import httpx
import io

c = httpx.Client(follow_redirects=True, timeout=30, headers={"User-Agent": "Mozilla/5.0"})

# 1. Download and inspect CC permit report Jan 2026
print("=== CC PERMIT REPORT JAN 2026 (XLSX) ===")
r = c.get("https://www.corpuschristitx.gov/media/ba1nbcm4/permit-report-jan-2026-excel-version.xlsx")
print(f"Status: {r.status_code}, len={len(r.content)}")
if r.status_code == 200:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  Sheet: {sn} ({ws.max_row} rows x {ws.max_column} cols)")
        # Print first 5 rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True), 1):
            print(f"    Row {row_idx}: {list(row)[:12]}")

# 2. Download CC CO report Jan 2026
print("\n=== CC CO REPORT JAN 2026 (XLSX) ===")
r = c.get("https://www.corpuschristitx.gov/media/riplqktg/coo-report-jan-2026-excel-version.xlsx")
print(f"Status: {r.status_code}, len={len(r.content)}")
if r.status_code == 200:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  Sheet: {sn} ({ws.max_row} rows x {ws.max_column} cols)")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True), 1):
            print(f"    Row {row_idx}: {list(row)[:12]}")

c.close()
print("\nDone.")
